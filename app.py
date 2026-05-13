import streamlit as st
import pandas as pd
from datetime import date, datetime
import io
import json
from utils import (
    load_projects, save_project, delete_project, duplicate_project,
    parse_project, load_shops_db, save_shop_db, delete_shop_db,
    load_templates, save_template, delete_template,
    calc_group, new_group, new_item,
    STATUS_OPTIONS, STATUS_COLORS, STATUS_TEXT, ITEM_CATEGORIES
)

st.set_page_config(page_title="Price Comparison", page_icon="💼", layout="wide")

st.markdown("""
<style>
/* ===== Global ===== */
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
html, body, [class*="css"] { font-family: 'Inter', sans-serif; }
.block-container { padding: 1.5rem 2rem 2rem; max-width: 1400px; }

/* ===== Sidebar ===== */
[data-testid="stSidebar"] {
    background: linear-gradient(160deg, #0f2027, #203a43, #2c5364);
    border-right: none;
}
[data-testid="stSidebar"] * { color: #e2e8f0 !important; }
[data-testid="stSidebar"] .stRadio label {
    background: rgba(255,255,255,0.07);
    border-radius: 10px;
    padding: 10px 14px;
    margin-bottom: 6px;
    display: block;
    cursor: pointer;
    transition: background 0.2s;
    font-size: 14px;
}
[data-testid="stSidebar"] .stRadio label:hover { background: rgba(255,255,255,0.15); }
[data-testid="stSidebar"] hr { border-color: rgba(255,255,255,0.15) !important; }

/* ===== Page header ===== */
.page-header {
    display: flex; align-items: center; gap: 14px;
    padding: 0 0 1.2rem; border-bottom: 2px solid #e2e8f0; margin-bottom: 1.5rem;
}
.page-header-icon {
    width: 44px; height: 44px; border-radius: 12px;
    background: linear-gradient(135deg, #1D9E75, #0F6E56);
    display: flex; align-items: center; justify-content: center;
    font-size: 20px; flex-shrink: 0;
}
.page-header-title { font-size: 22px; font-weight: 700; color: #0f172a; margin: 0; }
.page-header-sub { font-size: 13px; color: #64748b; margin: 2px 0 0; }

/* ===== Cards ===== */
.card {
    background: white; border: 1px solid #e2e8f0;
    border-radius: 14px; padding: 1.25rem 1.5rem;
    margin-bottom: 1rem; box-shadow: 0 1px 3px rgba(0,0,0,0.04);
    transition: box-shadow 0.2s;
}
.card:hover { box-shadow: 0 4px 12px rgba(0,0,0,0.08); }

/* ===== Project list item ===== */
.proj-card {
    background: white; border: 1px solid #e2e8f0; border-radius: 14px;
    padding: 1.1rem 1.4rem; margin-bottom: 10px;
    display: flex; align-items: center; gap: 16px;
    transition: all 0.2s; cursor: default;
}
.proj-card:hover { border-color: #1D9E75; box-shadow: 0 2px 12px rgba(29,158,117,0.1); }
.proj-icon {
    width: 42px; height: 42px; border-radius: 10px;
    background: #E1F5EE; display: flex; align-items: center;
    justify-content: center; font-size: 18px; flex-shrink: 0;
}
.proj-title { font-size: 15px; font-weight: 600; color: #0f172a; margin: 0; }
.proj-meta { font-size: 12px; color: #94a3b8; margin: 3px 0 0; }

/* ===== Status badge ===== */
.badge {
    display: inline-block; font-size: 11px; font-weight: 600;
    padding: 3px 10px; border-radius: 20px; white-space: nowrap;
}

/* ===== Section title ===== */
.sec-title {
    font-size: 13px; font-weight: 600; color: #64748b;
    text-transform: uppercase; letter-spacing: .06em;
    margin: 1.2rem 0 .6rem; display: flex; align-items: center; gap: 6px;
}

/* ===== Metric cards ===== */
.metric-row { display: flex; gap: 12px; margin-bottom: 1rem; flex-wrap: wrap; }
.metric-card {
    flex: 1; min-width: 130px; background: white; border: 1px solid #e2e8f0;
    border-radius: 12px; padding: 14px 16px;
}
.metric-label { font-size: 11px; color: #94a3b8; font-weight: 500; text-transform: uppercase; letter-spacing: .05em; }
.metric-value { font-size: 22px; font-weight: 700; color: #0f172a; margin-top: 4px; }
.metric-sub { font-size: 11px; color: #64748b; margin-top: 2px; }

/* ===== Winner highlight ===== */
.winner-card { border-left: 3px solid #1D9E75 !important; }
.winner-label { color: #1D9E75 !important; }

/* ===== Form styles ===== */
.stTextInput input, .stNumberInput input, .stSelectbox select,
.stTextArea textarea {
    border-radius: 8px !important; border: 1.5px solid #e2e8f0 !important;
    font-size: 14px !important; transition: border-color .2s !important;
}
.stTextInput input:focus, .stNumberInput input:focus { border-color: #1D9E75 !important; }

/* ===== Buttons ===== */
.stButton > button {
    border-radius: 8px !important; font-size: 13px !important;
    font-weight: 500 !important; transition: all .15s !important;
    border: 1.5px solid #e2e8f0 !important;
}
.stButton > button[kind="primary"] {
    background: linear-gradient(135deg, #1D9E75, #0F6E56) !important;
    color: white !important; border-color: transparent !important;
}
.stButton > button[kind="primary"]:hover { opacity: 0.9 !important; transform: translateY(-1px); }

/* ===== Tabs ===== */
.stTabs [data-baseweb="tab-list"] {
    gap: 4px; background: #f8fafc; border-radius: 10px;
    padding: 4px; border: none;
}
.stTabs [data-baseweb="tab"] {
    border-radius: 8px !important; font-size: 13px !important;
    font-weight: 500 !important; padding: 6px 16px !important;
    border: none !important; background: transparent !important;
}
.stTabs [aria-selected="true"] {
    background: white !important; color: #1D9E75 !important;
    box-shadow: 0 1px 4px rgba(0,0,0,0.1) !important;
}

/* ===== Table ===== */
.stDataFrame { border-radius: 12px !important; border: 1px solid #e2e8f0 !important; }

/* ===== Expander ===== */
.streamlit-expanderHeader {
    background: #f8fafc !important; border-radius: 10px !important;
    font-weight: 500 !important; font-size: 14px !important;
}

/* ===== Divider ===== */
hr { border-color: #f1f5f9 !important; }

/* ===== User info in sidebar ===== */
.user-chip {
    background: rgba(255,255,255,0.12); border-radius: 10px;
    padding: 10px 14px; margin-bottom: 12px;
    display: flex; align-items: center; gap: 10px;
}
.user-avatar {
    width: 32px; height: 32px; border-radius: 50%;
    background: linear-gradient(135deg,#1D9E75,#0F6E56);
    display: flex; align-items: center; justify-content: center;
    font-size: 13px; font-weight: 700; color: white; flex-shrink: 0;
}
.user-name { font-size: 13px; font-weight: 600; }
.user-role { font-size: 11px; opacity: .7; margin-top: 1px; }
</style>
""", unsafe_allow_html=True)

# ===== LOGIN =====
for k,v in [("logged_in",False),("username",""),("display_name","")]:
    if k not in st.session_state: st.session_state[k] = v

if not st.session_state["logged_in"]:
    # Login page
    st.markdown("""
    <div style="display:flex;flex-direction:column;align-items:center;justify-content:center;min-height:80vh;gap:0">
      <div style="text-align:center;margin-bottom:2rem">
        <div style="width:64px;height:64px;border-radius:18px;background:linear-gradient(135deg,#1D9E75,#0F6E56);
             display:flex;align-items:center;justify-content:center;font-size:28px;margin:0 auto 16px">💼</div>
        <h1 style="font-size:26px;font-weight:700;color:#0f172a;margin:0">ระบบเปรียบเทียบราคา</h1>
        <p style="color:#64748b;font-size:14px;margin:6px 0 0">Price Comparison System</p>
      </div>
    </div>
    """, unsafe_allow_html=True)
    col = st.columns([1,1,1])[1]
    with col:
        with st.form("login_form"):
            st.markdown("**เข้าสู่ระบบ**")
            username = st.text_input("Username", placeholder="กรอก username")
            password = st.text_input("Password", type="password", placeholder="กรอก password")
            submitted = st.form_submit_button("เข้าสู่ระบบ", type="primary", use_container_width=True)
            if submitted:
                users = st.secrets.get("users",{})
                if username in users and users[username]["password"] == password:
                    st.session_state.update({"logged_in":True,"username":username,
                                             "display_name":users[username]["name"]})
                    st.rerun()
                else:
                    st.error("Username หรือ Password ไม่ถูกต้อง")
    st.stop()

# ===== SIDEBAR =====
with st.sidebar:
    initials = st.session_state["display_name"][:2].upper() if st.session_state["display_name"] else "U"
    st.markdown(f"""
    <div class="user-chip">
      <div class="user-avatar">{initials}</div>
      <div>
        <div class="user-name">{st.session_state["display_name"]}</div>
        <div class="user-role">ผู้ใช้งาน</div>
      </div>
    </div>
    """, unsafe_allow_html=True)

    menu = st.radio("เมนู", ["📋  โครงการ","🏪  ร้านค้า","📦  Templates"],
                    label_visibility="collapsed")
    st.markdown("---")
    if st.button("ออกจากระบบ", use_container_width=True):
        for k in list(st.session_state.keys()): del st.session_state[k]
        st.rerun()

# ===== SESSION DEFAULTS =====
for k,v in [("mode","list"),("current_project_id",None),
            ("shops",["บริษัท A","บริษัท B","บริษัท C"]),
            ("groups",[{"name":"กลุ่มที่ 1","items":[],"discounts":{}}]),
            ("doc_title","เปรียบเทียบราคา"),("project_name",""),
            ("vat_rate",7.0),("shop_discounts",{}),("doc_date",date.today()),
            ("tags",[]),("status","กำลังดำเนินการ")]:
    if k not in st.session_state: st.session_state[k] = v

def load_into_session(r):
    shops, shop_discounts, groups = parse_project(r)
    try: tags = json.loads(r.get("tags","[]"))
    except: tags = []
    st.session_state.update({
        "doc_title":r.get("title",""), "project_name":r.get("project_name",""),
        "vat_rate":float(r.get("vat_rate",7.0)), "shops":shops,
        "shop_discounts":shop_discounts, "groups":groups,
        "tags":tags, "status":r.get("status","กำลังดำเนินการ"),
    })
    try: st.session_state["doc_date"] = datetime.strptime(r.get("date",""),"%Y-%m-%d").date()
    except: st.session_state["doc_date"] = date.today()

def do_save(show_msg=True):
    pid = save_project({
        "title":st.session_state["doc_title"],
        "project_name":st.session_state["project_name"],
        "date":st.session_state["doc_date"].strftime("%Y-%m-%d"),
        "vat_rate":st.session_state["vat_rate"],
        "shops":st.session_state["shops"],
        "shop_discounts":st.session_state["shop_discounts"],
        "groups":st.session_state["groups"],
        "tags":st.session_state["tags"],
        "status":st.session_state["status"],
    }, st.session_state["display_name"], st.session_state["current_project_id"])
    st.session_state.update({"current_project_id":pid,"mode":"edit"})
    if show_msg: st.success("✅ บันทึกเรียบร้อย!")

# ============================================================
# MENU: โครงการ
# ============================================================
if "📋" in menu:

    # ===== LIST =====
    if st.session_state["mode"] == "list":
        st.markdown("""
        <div class="page-header">
          <div class="page-header-icon">📋</div>
          <div>
            <div class="page-header-title">รายการโครงการ</div>
            <div class="page-header-sub">จัดการและเปรียบเทียบราคาทุกโครงการ</div>
          </div>
        </div>
        """, unsafe_allow_html=True)

        c1,c2 = st.columns([3,1])
        with c2:
            if st.button("➕ สร้างโครงการใหม่", type="primary", use_container_width=True):
                st.session_state.update({
                    "mode":"new","current_project_id":None,
                    "shops":["บริษัท A","บริษัท B","บริษัท C"],
                    "groups":[{"name":"กลุ่มที่ 1","items":[],"discounts":{}}],
                    "doc_title":"เปรียบเทียบราคา","project_name":"",
                    "vat_rate":7.0,"shop_discounts":{},"doc_date":date.today(),
                    "tags":[],"status":"กำลังดำเนินการ",
                })
                st.rerun()

        # Search & filter
        with c1:
            sc1,sc2,sc3 = st.columns([2,1,1])
            search_q      = sc1.text_input("", placeholder="🔍 ค้นหาโครงการ...", label_visibility="collapsed")
            filter_status = sc2.selectbox("", ["ทั้งหมด"]+STATUS_OPTIONS, label_visibility="collapsed")
            sort_by       = sc3.selectbox("", ["ใหม่สุด","เก่าสุด","ชื่อ A-Z"], label_visibility="collapsed")

        with st.spinner(""):
            projects = load_projects()

        if search_q:
            projects = [p for p in projects if search_q.lower() in p.get("title","").lower()
                        or search_q.lower() in p.get("project_name","").lower()]
        if filter_status != "ทั้งหมด":
            projects = [p for p in projects if p.get("status","") == filter_status]
        if sort_by == "ใหม่สุด": projects = list(reversed(projects))
        elif sort_by == "ชื่อ A-Z": projects = sorted(projects, key=lambda x: x.get("title",""))

        if not projects:
            st.markdown("""
            <div style="text-align:center;padding:4rem 2rem;color:#94a3b8">
              <div style="font-size:48px;margin-bottom:12px">📂</div>
              <div style="font-size:16px;font-weight:500;color:#64748b">ยังไม่มีโครงการ</div>
              <div style="font-size:13px;margin-top:6px">กด 'สร้างโครงการใหม่' เพื่อเริ่มต้น</div>
            </div>
            """, unsafe_allow_html=True)
        else:
            for r in projects:
                _, _, groups_r = parse_project(r)
                total_items = sum(len(g.get("items",[])) for g in groups_r)
                try: tags = json.loads(r.get("tags","[]"))
                except: tags = []
                status = r.get("status","กำลังดำเนินการ")
                sbg  = STATUS_COLORS.get(status,"#f1f5f9")
                stxt = STATUS_TEXT.get(status,"#1e293b")

                with st.container():
                    ca,cb = st.columns([5,2])
                    with ca:
                        tag_html = "".join(f'<span style="background:#f1f5f9;color:#64748b;font-size:11px;padding:2px 8px;border-radius:20px;margin-left:4px">{t}</span>' for t in tags)
                        st.markdown(f"""
                        <div style="display:flex;align-items:flex-start;gap:14px;padding:14px 0">
                          <div style="width:42px;height:42px;border-radius:10px;background:#E1F5EE;
                               display:flex;align-items:center;justify-content:center;font-size:18px;flex-shrink:0">📁</div>
                          <div>
                            <div style="font-size:15px;font-weight:600;color:#0f172a">
                              {r.get('title','')}
                              <span class="badge" style="background:{sbg};color:{stxt};margin-left:8px">{status}</span>
                              {tag_html}
                            </div>
                            <div style="font-size:12px;color:#94a3b8;margin-top:4px">
                              📁 {r.get('project_name','')} &nbsp;·&nbsp; 🗓 {r.get('date','')} &nbsp;·&nbsp; 👤 {r.get('created_by','')} &nbsp;·&nbsp; {len(groups_r)} กลุ่ม · {total_items} รายการ
                            </div>
                          </div>
                        </div>
                        """, unsafe_allow_html=True)
                    with cb:
                        st.markdown("<div style='height:14px'></div>", unsafe_allow_html=True)
                        ba,bb,bc = st.columns(3)
                        if ba.button("แก้ไข", key=f"e_{r['id']}", use_container_width=True):
                            load_into_session(r)
                            st.session_state.update({"mode":"edit","current_project_id":r["id"]})
                            st.rerun()
                        if bb.button("สำเนา", key=f"dup_{r['id']}", use_container_width=True):
                            duplicate_project(r["id"], st.session_state["display_name"])
                            st.rerun()
                        if bc.button("ลบ", key=f"d_{r['id']}", use_container_width=True):
                            delete_project(r["id"]); st.rerun()
                    st.markdown("<hr style='margin:0;border-color:#f1f5f9'>", unsafe_allow_html=True)

    # ===== NEW / EDIT =====
    elif st.session_state["mode"] in ["new","edit"]:
        is_edit = st.session_state["mode"] == "edit"

        # Top bar
        c1,c2,c3 = st.columns([4,1,1])
        c1.markdown(f"""
        <div style="display:flex;align-items:center;gap:10px;padding:4px 0">
          <span style="font-size:20px">{'✏️' if is_edit else '➕'}</span>
          <span style="font-size:20px;font-weight:700;color:#0f172a">{'แก้ไขโครงการ' if is_edit else 'สร้างโครงการใหม่'}</span>
        </div>
        """, unsafe_allow_html=True)
        if c2.button("💾 บันทึก", type="primary", use_container_width=True):
            do_save(); st.rerun()
        if c3.button("← กลับ", use_container_width=True):
            st.session_state["mode"] = "list"; st.rerun()

        st.markdown("<hr style='margin:.5rem 0 1rem'>", unsafe_allow_html=True)

        # Sidebar settings
        with st.sidebar:
            st.markdown("<div class='sec-title'>⚙️ ตั้งค่า</div>", unsafe_allow_html=True)
            st.session_state["doc_title"]    = st.text_input("ชื่อเอกสาร",  st.session_state["doc_title"])
            st.session_state["project_name"] = st.text_input("ชื่อโครงการ", st.session_state["project_name"])
            st.session_state["doc_date"]     = st.date_input("วันที่", value=st.session_state["doc_date"])
            c_vat, c_status = st.columns(2)
            st.session_state["vat_rate"] = c_vat.number_input("VAT %", value=st.session_state["vat_rate"],
                                                                min_value=0.0, max_value=30.0, step=0.5)
            st.session_state["status"]   = c_status.selectbox("สถานะ", STATUS_OPTIONS,
                                                               index=STATUS_OPTIONS.index(st.session_state["status"])
                                                               if st.session_state["status"] in STATUS_OPTIONS else 0)
            tag_input = st.text_input("Tags", ", ".join(st.session_state["tags"]),
                                      placeholder="เช่น IT, ก่อสร้าง")
            st.session_state["tags"] = [t.strip() for t in tag_input.split(",") if t.strip()]

            st.markdown("<div class='sec-title'>🏪 ร้านค้า</div>", unsafe_allow_html=True)
            shops_db = load_shops_db()
            if shops_db:
                shop_names_db = [s["name"] for s in shops_db]
                selected_db = st.multiselect("เลือกจากฐานข้อมูล", shop_names_db, label_visibility="collapsed",
                                             placeholder="เลือกร้านจากฐานข้อมูล...")
                if selected_db and st.button("+ เพิ่มร้านที่เลือก", use_container_width=True):
                    for sn in selected_db:
                        if sn not in st.session_state["shops"]:
                            st.session_state["shops"].append(sn)
                    st.rerun()

            shops_to_delete = None
            for i, s in enumerate(st.session_state["shops"]):
                col_in, col_del = st.columns([4, 0.7])
                st.session_state["shops"][i] = col_in.text_input(f"ร้านที่ {i+1}", s, key=f"shop_name_{i}")
                if len(st.session_state["shops"]) > 2:
                    if col_del.button("🗑", key=f"del_shop_btn_{i}", help=f"ลบร้านที่ {i+1}"):
                        shops_to_delete = i
            if shops_to_delete is not None:
                st.session_state["shops"].pop(shops_to_delete)
                for j in range(len(st.session_state["shops"]), len(st.session_state["shops"])+5):
                    if f"shop_name_{j}" in st.session_state: del st.session_state[f"shop_name_{j}"]
                st.rerun()
            if st.button("➕ เพิ่มร้านค้า", use_container_width=True):
                st.session_state["shops"].append(f"บริษัท {chr(65+len(st.session_state['shops']))}"); st.rerun()

            st.markdown("<div class='sec-title'>🎁 Special Discount (฿)</div>", unsafe_allow_html=True)
            for g_idx, grp in enumerate(st.session_state["groups"]):
                if "discounts" not in grp: grp["discounts"] = {}
                if len(st.session_state["groups"]) > 1:
                    st.caption(grp["name"])
                for s in st.session_state["shops"]:
                    cur = float(grp["discounts"].get(s,0.0))
                    grp["discounts"][s] = st.number_input(s, value=cur, min_value=0.0, step=1.0,
                                                           key=f"disc_{g_idx}_{s}")

        shops = st.session_state["shops"]
        vat   = st.session_state["vat_rate"]

        tab1, tab2 = st.tabs(["✏️  กรอกข้อมูล", "📊  ตารางเปรียบเทียบ"])

        # ===== TAB 1 =====
        with tab1:
            col_add, col_tmpl = st.columns([1,3])
            if col_add.button("➕ เพิ่มกลุ่มสินค้า"):
                st.session_state["groups"].append(new_group(len(st.session_state["groups"])+1)); st.rerun()

            templates = load_templates()
            if templates:
                tmpl_names = ["-- เลือก Template --"] + [t["name"] for t in templates]
                sel_tmpl = col_tmpl.selectbox("โหลด Template", tmpl_names, label_visibility="collapsed")
                if sel_tmpl != "-- เลือก Template --":
                    tmpl = next((t for t in templates if t["name"]==sel_tmpl), None)
                    if tmpl and st.button(f"➕ โหลด '{sel_tmpl}'"):
                        try:
                            for it in json.loads(tmpl["items"]):
                                ni = new_item(shops)
                                ni.update({"name":it.get("name",""),"unit":it.get("unit","set"),
                                           "qty":float(it.get("qty",1)),"category":it.get("category","วัสดุ/อุปกรณ์")})
                                st.session_state["groups"][0]["items"].append(ni)
                            st.rerun()
                        except: pass

            group_names = [f"{i+1}. {g['name']}" for i,g in enumerate(st.session_state["groups"])]
            sub_tabs = st.tabs(group_names)

            for g_idx, sub_tab in enumerate(sub_tabs):
                with sub_tab:
                    grp = st.session_state["groups"][g_idx]
                    if "items" not in grp: grp["items"] = []
                    if "discounts" not in grp: grp["discounts"] = {}

                    gc1,gc2,gc3 = st.columns([3,1,1])
                    grp["name"] = gc1.text_input("ชื่อกลุ่ม", grp["name"], key=f"gname_{g_idx}")
                    if gc2.button("➕ เพิ่มสินค้า", key=f"gadd_{g_idx}", use_container_width=True):
                        grp["items"].append(new_item(shops)); st.rerun()
                    if gc3.button("🗑 ลบกลุ่ม", key=f"gdel_{g_idx}", use_container_width=True) and len(st.session_state["groups"]) > 1:
                        st.session_state["groups"].pop(g_idx); st.rerun()

                    if grp["items"] and st.button(f"💾 บันทึกกลุ่มนี้เป็น Template", key=f"save_tmpl_{g_idx}"):
                        save_template({
                            "name":f"{grp['name']} ({datetime.now().strftime('%d/%m/%y')})",
                            "category":"ทั่วไป",
                            "items":[{"name":it["name"],"unit":it["unit"],"qty":it["qty"],"category":it.get("category","")} for it in grp["items"]]
                        }, st.session_state["display_name"])
                        st.success("บันทึก Template แล้ว!")

                    n = len(grp["items"])
                    if n == 0:
                        st.markdown("""
                        <div style="text-align:center;padding:2rem;color:#94a3b8;background:#f8fafc;border-radius:12px;margin:12px 0">
                          กด 'เพิ่มสินค้า' เพื่อเพิ่มรายการในกลุ่มนี้
                        </div>
                        """, unsafe_allow_html=True)
                    else:
                        to_del = None
                        for idx in range(n):
                            it = grp["items"][idx]
                            for key,default in [("prices",{}),("item_discounts",{}),("notes",{})]:
                                if key not in it: it[key] = {}
                            for s in shops:
                                for key,default in [("prices",0.0),("item_discounts",0.0),("notes","")]:
                                    if s not in it[key]: it[key][s] = default

                            with st.expander(f"{idx+1}. {it['name']}  ·  {it.get('category','')}", expanded=True):
                                ca,cb,cc,cd,ce = st.columns([2.5,1,0.8,1.5,0.5])
                                it["name"]     = ca.text_input("ชื่อสินค้า/บริการ", it["name"],  key=f"n_{g_idx}_{idx}")
                                it["unit"]     = cb.text_input("หน่วย",             it["unit"],  key=f"u_{g_idx}_{idx}")
                                it["qty"]      = cc.number_input("จำนวน", value=float(it["qty"]), min_value=0.0, step=1.0, key=f"q_{g_idx}_{idx}")
                                it["category"] = cd.selectbox("ประเภท", ITEM_CATEGORIES,
                                                              index=ITEM_CATEGORIES.index(it.get("category","วัสดุ/อุปกรณ์"))
                                                              if it.get("category") in ITEM_CATEGORIES else 0,
                                                              key=f"cat_{g_idx}_{idx}")
                                if ce.button("🗑", key=f"d_{g_idx}_{idx}"): to_del = idx

                                st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)
                                hcols = st.columns(len(shops))
                                for si,s in enumerate(shops):
                                    hcols[si].markdown(f"<div style='font-size:12px;font-weight:600;color:#1D9E75;margin-bottom:4px'>{s}</div>", unsafe_allow_html=True)

                                pcols = st.columns(len(shops))
                                for si,s in enumerate(shops):
                                    it["prices"][s] = pcols[si].number_input("ราคา/หน่วย (฿)", value=float(it["prices"].get(s,0)),
                                                                               min_value=0.0, step=1.0, key=f"p_{g_idx}_{idx}_{si}")
                                dcols = st.columns(len(shops))
                                for si,s in enumerate(shops):
                                    it["item_discounts"][s] = dcols[si].number_input("ส่วนลด/หน่วย (฿)", value=float(it["item_discounts"].get(s,0)),
                                                                                      min_value=0.0, step=1.0, key=f"id_{g_idx}_{idx}_{si}")
                                st.markdown("**หมายเหตุ**")
                                ncols = st.columns(len(shops))
                                for si,s in enumerate(shops):
                                    it["notes"][s] = ncols[si].text_area(f"({s})", value=it["notes"].get(s,""),
                                                                          placeholder="เช่น รับประกัน 1 ปี",
                                                                          height=60, key=f"note_{g_idx}_{idx}_{si}",
                                                                          label_visibility="collapsed")
                                st.markdown("---")
                                scols = st.columns(len(shops))
                                qty = float(it["qty"])
                                for si,s in enumerate(shops):
                                    price = float(it["prices"].get(s,0))
                                    disc  = float(it["item_discounts"].get(s,0))
                                    sub   = max(price-disc,0)*qty
                                    disc_txt = f" (ลด ฿{disc:,.0f}/หน่วย)" if disc > 0 else ""
                                    scols[si].markdown(f"<div style='font-size:13px'>ยอดรวม{disc_txt}<br><b style='font-size:16px'>฿{sub:,.2f}</b></div>", unsafe_allow_html=True)

                        if to_del is not None:
                            grp["items"].pop(to_del); st.rerun()

            st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)
            if st.button("💾 บันทึกลง Google Sheets", type="primary", use_container_width=False):
                do_save(); st.rerun()

        # ===== TAB 2 =====
        with tab2:
            doc_date = st.session_state["doc_date"]
            groups   = st.session_state["groups"]

            if not any(g.get("items") for g in groups):
                st.info("ยังไม่มีข้อมูล — กรอกข้อมูลในแท็บ 'กรอกข้อมูล' ก่อนครับ")
            else:
                all_totals = {s:0.0 for s in shops}
                for grp in groups:
                    if not grp.get("items"): continue
                    _,_,_,_,gt = calc_group(grp["items"],shops,vat,grp.get("discounts",{}))
                    for s in shops: all_totals[s] += gt[s]

                data_groups = [(i,g) for i,g in enumerate(groups) if g.get("items")]
                tab_labels  = [f"{i+1}. {g['name']}" for i,g in data_groups]
                if len(data_groups) > 1: tab_labels.append("📊 สรุปรวม")
                sub_tabs2 = st.tabs(tab_labels)

                for tab_i,(g_idx,grp) in enumerate(data_groups):
                    with sub_tabs2[tab_i]:
                        items_data = grp.get("items",[])
                        shop_disc  = grp.get("discounts",{})
                        grand_sub,grand_disc,grand_after_disc,grand_vat,grand_total = calc_group(
                            items_data,shops,vat,shop_disc)

                        valid = {s:grand_total[s] for s in shops if grand_sub[s]>0}
                        best  = min(valid,key=valid.get) if valid else None

                        if valid and best:
                            save_amt = max(valid.values())-min(valid.values())
                            cols = st.columns(4)
                            cols[0].metric("ร้านถูกสุด (หลังส่วนลด)", best)
                            cols[1].metric("ยอดสุทธิถูกสุด", f"฿{valid[best]:,.2f}")
                            cols[2].metric("ประหยัดได้", f"฿{save_amt:,.2f}")
                            cols[3].metric("รายการ", len(items_data))

                        st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

                        # ตาราง แยกตามประเภท
                        categories = list(dict.fromkeys(it.get("category","วัสดุ/อุปกรณ์") for it in items_data))
                        for cat in categories:
                            cat_items = [it for it in items_data if it.get("category","วัสดุ/อุปกรณ์")==cat]
                            st.markdown(f"<div style='font-size:12px;font-weight:600;color:#64748b;text-transform:uppercase;letter-spacing:.05em;margin:12px 0 6px'>{cat}</div>", unsafe_allow_html=True)
                            rows = []
                            for i,it in enumerate(cat_items):
                                qty = float(it["qty"])
                                row = {"#":i+1,"รายการ":it["name"],"จำนวน":int(qty),"หน่วย":it["unit"]}
                                tot_vals = []
                                for s in shops:
                                    price = float(it["prices"].get(s,0))
                                    disc  = float(it["item_discounts"].get(s,0))
                                    total = max(price-disc,0)*qty
                                    row[f"{s}\nUnit Price"] = price
                                    row[f"{s}\nส่วนลด(฿)"] = disc if disc > 0 else ""
                                    row[f"{s}\nTotal"]      = round(total,2)
                                    tot_vals.append(total)
                                valid_tv = [v for v in tot_vals if v>0]
                                if valid_tv: row["ถูกสุด"] = shops[tot_vals.index(min(valid_tv))]
                                rows.append(row)

                            df = pd.DataFrame(rows)
                            def hi(row):
                                tc=[c for c in row.index if "Total" in str(c)]
                                styles=[""]*len(row)
                                vals={c:float(row[c]) for c in tc if str(row[c]) not in ["","0.0","0"]}
                                if not vals: return styles
                                mn=min(vals.values())
                                for i2,col in enumerate(row.index):
                                    if col in tc and str(row[col]) not in ["","0"] and float(row[col])==mn:
                                        styles[i2]="background-color:#E1F5EE;font-weight:600;color:#0F6E56"
                                    elif col in tc and str(row[col]) not in ["","0"] and float(row[col])>mn:
                                        styles[i2]="color:#dc2626"
                                return styles

                            def fmt_disc(v):
                                if v=="" or v is None: return ""
                                try: return f"฿{float(v):,.2f}"
                                except: return ""

                            fmt = {c:"฿{:,.2f}" for c in df.columns if "Total" in str(c) or "Price" in str(c)}
                            disc_cols = [c for c in df.columns if "ส่วนลด" in str(c)]
                            styled = df.style.apply(hi,axis=1).format(fmt)
                            for dc in disc_cols: styled = styled.format(fmt_disc, subset=[dc])
                            st.dataframe(styled, use_container_width=True, hide_index=True)

                        # สรุปยอด
                        st.markdown("<div style='margin-top:16px'></div>", unsafe_allow_html=True)
                        sum_rows = [
                            ("ยอดรวมก่อนส่วนลด",   grand_sub,        False),
                            ("SPECIAL DISCOUNT (-)", grand_disc,       True),
                            ("TOTAL (EXC. VAT)",    grand_after_disc, False),
                            (f"VAT {vat:.0f}%",     grand_vat,        False),
                            ("TOTAL (INC. VAT)",    grand_total,      False),
                        ]
                        hcols = st.columns([2]+[1]*len(shops))
                        hcols[0].markdown("<div style='font-size:12px;font-weight:600;color:#64748b'>รายการ</div>", unsafe_allow_html=True)
                        for si,s in enumerate(shops):
                            lbl = "🏆 " if valid and s==best else ""
                            hcols[si+1].markdown(f"<div style='font-size:12px;font-weight:600;color:{'#1D9E75' if valid and s==best else '#0f172a'}'>{lbl}{s}</div>", unsafe_allow_html=True)

                        for label,vals,is_disc in sum_rows:
                            is_total = "TOTAL (INC." in label
                            rcols = st.columns([2]+[1]*len(shops))
                            style = "font-weight:600;color:#0f172a" if is_total else "color:#374151"
                            rcols[0].markdown(f"<div style='font-size:13px;{style};padding:4px 0'>{label}</div>", unsafe_allow_html=True)
                            for si,s in enumerate(shops):
                                v = vals.get(s,0)
                                is_b = is_total and valid and s==best
                                disp = f"-฿{v:,.2f}" if is_disc else f"฿{v:,.2f}"
                                color = "#1D9E75" if is_b else ("#374151" if not is_total else "#0f172a")
                                bg = "background:#E1F5EE;border-radius:6px;padding:2px 8px;" if is_b else ""
                                rcols[si+1].markdown(f"<div style='font-size:13px;{style};{bg}color:{color};padding:4px 0'>{disp}</div>", unsafe_allow_html=True)

                        # หมายเหตุ
                        has_notes = any(it.get("notes",{}).get(s,"").strip() for it in items_data for s in shops)
                        if has_notes:
                            st.markdown("---")
                            st.markdown("<div style='font-size:12px;font-weight:600;color:#64748b;margin-bottom:8px'>หมายเหตุ</div>", unsafe_allow_html=True)
                            ncols = st.columns(len(shops))
                            for si,s in enumerate(shops):
                                lbl = "🏆 " if valid and s==best else ""
                                with ncols[si]:
                                    st.markdown(f"**{lbl}{s}**")
                                    for it in items_data:
                                        note = it.get("notes",{}).get(s,"").strip()
                                        if note:
                                            st.markdown(f"**· {it['name']}**")
                                            st.markdown(f"<div style='font-size:12px;color:#64748b'>{note}</div>", unsafe_allow_html=True)

                if len(data_groups) > 1:
                    with sub_tabs2[-1]:
                        valid_all = {s:all_totals[s] for s in shops if all_totals[s]>0}
                        if valid_all:
                            best_all  = min(valid_all,key=valid_all.get)
                            save_all  = max(valid_all.values())-min(valid_all.values())
                            m1,m2,m3  = st.columns(3)
                            m1.metric("ร้านถูกสุดโดยรวม", best_all)
                            m2.metric("ยอดรวมถูกสุด",     f"฿{valid_all[best_all]:,.2f}")
                            m3.metric("ประหยัดได้รวม",     f"฿{save_all:,.2f}")

                        sum_rows2 = []
                        for g_idx2,grp2 in enumerate(groups):
                            if not grp2.get("items"): continue
                            _,_,_,_,gt = calc_group(grp2["items"],shops,vat,grp2.get("discounts",{}))
                            row2 = {"กลุ่ม":f"กลุ่ม {g_idx2+1}: {grp2['name']}"}
                            for s in shops: row2[f"{s} TOTAL"] = round(gt[s],2)
                            sum_rows2.append(row2)
                        total_row2 = {"กลุ่ม":"รวมทั้งหมด"}
                        for s in shops: total_row2[f"{s} TOTAL"] = round(all_totals[s],2)
                        sum_rows2.append(total_row2)
                        df2 = pd.DataFrame(sum_rows2)
                        def hi2(row):
                            tc=[c for c in row.index if "TOTAL" in str(c)]
                            styles=[""]*len(row)
                            vals={c:float(row[c]) for c in tc if float(row[c])>0}
                            if not vals: return styles
                            mn=min(vals.values())
                            for i3,col in enumerate(row.index):
                                if col in tc and float(row[col])==mn and float(row[col])>0:
                                    styles[i3]="background-color:#E1F5EE;font-weight:600;color:#0F6E56"
                                elif col in tc and float(row[col])>mn:
                                    styles[i3]="color:#dc2626"
                            return styles
                        fmt2 = {c:"฿{:,.2f}" for c in df2.columns if "TOTAL" in str(c)}
                        st.dataframe(df2.style.apply(hi2,axis=1).format(fmt2),use_container_width=True,hide_index=True)

                st.markdown("---")

                # ===== Export Excel =====
                def export_excel():
                    from openpyxl import Workbook
                    from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
                    from openpyxl.utils import get_column_letter
                    wb = Workbook(); ws = wb.active; ws.title = "Price Comparison"

                    # Color palette — เรียบง่าย ทันสมัย
                    C_HEADER   = "1E293B"   # header หลัก: slate-900
                    C_SUBHDR   = "334155"   # header ร้าน: slate-700
                    C_WIN      = "DCFCE7"   # winner: green-100
                    C_WIN_TXT  = "166534"   # winner text: green-800
                    C_DISC     = "FEF9C3"   # ส่วนลด: yellow-100
                    C_DISC_TXT = "854D0E"   # ส่วนลด text
                    C_SUMMARY  = "F8FAFC"   # summary bg: slate-50
                    C_TOTAL    = "E1F5EE"   # total row: green-50
                    C_TOTAL_TXT= "0F6E56"
                    C_STRIPE   = "F8FAFC"   # แถวคี่: slate-50
                    C_WHITE    = "FFFFFF"
                    C_TEXT     = "1E293B"
                    C_MUTED    = "64748B"
                    C_BORDER   = "E2E8F0"

                    thin   = Side(style="thin",   color=C_BORDER)
                    bdr    = Border(left=thin, right=thin, top=thin, bottom=thin)
                    bdr_none = Border()

                    def cs(r,c,val="",bold=False,color=C_TEXT,bg=None,align="left",
                           fmt=None,size=10,wrap=False,italic=False):
                        cell = ws.cell(row=r, column=c, value=val)
                        cell.font      = Font(bold=bold, color=color, size=size, name="Calibri", italic=italic)
                        if bg: cell.fill = PatternFill("solid", fgColor=bg)
                        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
                        if fmt: cell.number_format = fmt
                        cell.border = bdr
                        return cell

                    n_shops    = len(shops)
                    SHOP_START = 5
                    total_cols = SHOP_START + n_shops*2 - 1

                    # ===== แถว 1: ชื่อโครงการ =====
                    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
                    t = ws.cell(row=1, column=1,
                                value=f"เปรียบเทียบราคา  —  {st.session_state['project_name']}")
                    t.font      = Font(bold=True, size=15, color=C_HEADER, name="Calibri")
                    t.alignment = Alignment(horizontal="left", vertical="center")
                    t.border    = bdr_none
                    t.fill      = PatternFill("solid", fgColor=C_WHITE)
                    ws.row_dimensions[1].height = 28

                    # ===== แถว 2: รายละเอียด =====
                    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
                    d = ws.cell(row=2, column=1,
                                value=f"วันที่ {doc_date.strftime('%d/%m/%Y')}   |   VAT {vat:.0f}%   |   {st.session_state['doc_title']}")
                    d.font      = Font(size=9, color=C_MUTED, name="Calibri")
                    d.alignment = Alignment(horizontal="left", vertical="center")
                    d.border    = bdr_none
                    ws.row_dimensions[2].height = 16

                    # ===== แถว 3: เส้นคั่น =====
                    ws.row_dimensions[3].height = 6

                    current_row = 4

                    for g_idx, grp in enumerate(groups):
                        items_data = grp.get("items",[])
                        shop_disc  = grp.get("discounts",{})
                        if not items_data: continue

                        grand_sub,grand_disc_g,grand_after_disc,grand_vat,grand_total = calc_group(
                            items_data, shops, vat, shop_disc)
                        valid   = {s:grand_total[s] for s in shops if grand_sub[s]>0}
                        best_i  = shops.index(min(valid,key=valid.get)) if valid else None

                        # หัวกลุ่ม
                        ws.merge_cells(start_row=current_row, start_column=1,
                                       end_row=current_row, end_column=total_cols)
                        g = ws.cell(row=current_row, column=1,
                                    value=f"กลุ่ม {g_idx+1}  :  {grp['name']}")
                        g.font      = Font(bold=True, size=11, color=C_WHITE, name="Calibri")
                        g.fill      = PatternFill("solid", fgColor=C_HEADER)
                        g.alignment = Alignment(horizontal="left", vertical="center", indent=1)
                        g.border    = bdr_none
                        ws.row_dimensions[current_row].height = 22
                        current_row += 1

                        # header: Item / Detail / QTY / Unit
                        for c,label,w in [(1,"#",5),(2,"รายการ",36),(3,"จำนวน",8),(4,"หน่วย",8)]:
                            ws.merge_cells(start_row=current_row, start_column=c,
                                           end_row=current_row+1, end_column=c)
                            cell = cs(current_row, c, label, bold=True, color=C_WHITE,
                                      bg=C_SUBHDR, align="center", size=9)
                            ws.column_dimensions[get_column_letter(c)].width = w

                        # header: ชื่อร้าน (span 2 คอลัมน์)
                        for si, shop in enumerate(shops):
                            col_u = SHOP_START + si*2
                            col_t = col_u + 1
                            shop_bg = C_HEADER if si == best_i else C_SUBHDR
                            lbl_shop = ("★ " if si == best_i else "") + shop
                            ws.merge_cells(start_row=current_row, start_column=col_u,
                                           end_row=current_row, end_column=col_t)
                            cs(current_row, col_u, lbl_shop, bold=True,
                               color=C_WHITE, bg=shop_bg, align="center", size=9)
                            cs(current_row+1, col_u, "Unit Price", bold=True,
                               color="CBD5E1", bg=C_SUBHDR, align="center", size=8)
                            cs(current_row+1, col_t, "Total",      bold=True,
                               color="CBD5E1", bg=C_SUBHDR, align="center", size=8)
                            ws.column_dimensions[get_column_letter(col_u)].width = 13
                            ws.column_dimensions[get_column_letter(col_t)].width = 13

                        ws.row_dimensions[current_row].height   = 20
                        ws.row_dimensions[current_row+1].height = 16
                        current_row += 2

                        # data rows
                        for i, it in enumerate(items_data):
                            qty      = float(it["qty"])
                            row_bg   = C_WHITE if i%2==0 else C_STRIPE
                            raw_tots = [float(it["prices"].get(s,0))*qty for s in shops]
                            disc_tots= [float(it.get("item_discounts",{}).get(s,0))*qty for s in shops]
                            after_t  = [max(raw_tots[si]-disc_tots[si],0) for si in range(len(shops))]
                            valid_tv = [v for v in after_t if v>0]
                            best_p   = min(valid_tv) if valid_tv else None

                            cs(current_row, 1, i+1,          align="center", bg=row_bg, size=9)
                            cs(current_row, 2, f"[{it.get('category','')}] {it['name']}",
                               align="left", bg=row_bg, wrap=True, size=10, bold=True)
                            cs(current_row, 3, int(qty),      align="center", bg=row_bg, size=9)
                            cs(current_row, 4, it["unit"],    align="center", bg=row_bg, size=9)

                            for si, shop in enumerate(shops):
                                col_u = SHOP_START + si*2
                                col_t = col_u + 1
                                price = float(it["prices"].get(shop,0))
                                raw_t = raw_tots[si]
                                is_b  = best_p and after_t[si]==best_p and after_t[si]>0
                                cell_bg  = C_WIN   if is_b else row_bg
                                cell_txt = C_WIN_TXT if is_b else C_TEXT
                                cs(current_row, col_u, price, align="right", bg=cell_bg,
                                   fmt='#,##0.00', size=10, color=cell_txt)
                                cs(current_row, col_t, raw_t, align="right", bg=cell_bg,
                                   fmt='#,##0.00', size=10, bold=is_b, color=cell_txt)

                            ws.row_dimensions[current_row].height = 18
                            current_row += 1

                            # แถวส่วนลด (ถ้ามี)
                            has_disc = any(float(it.get("item_discounts",{}).get(s,0))>0 for s in shops)
                            if has_disc:
                                ws.merge_cells(start_row=current_row, start_column=1,
                                               end_row=current_row, end_column=4)
                                cs(current_row, 1, f"  ส่วนลด  {it['name']}", align="right",
                                   bg=C_DISC, size=8, color=C_DISC_TXT, italic=True)
                                for si, shop in enumerate(shops):
                                    col_u = SHOP_START + si*2
                                    col_t = col_u + 1
                                    dt = disc_tots[si]
                                    ws.merge_cells(start_row=current_row, start_column=col_u,
                                                   end_row=current_row, end_column=col_t)
                                    cs(current_row, col_u, -dt if dt>0 else "", align="right",
                                       bg=C_DISC, fmt='#,##0.00', size=8, color=C_DISC_TXT)
                                ws.row_dimensions[current_row].height = 14
                                current_row += 1

                        # Summary rows
                        sum_defs = [
                            ("SPECIAL DISCOUNT",  grand_disc_g,    C_DISC,    False, True,  C_DISC_TXT),
                            ("TOTAL (EXC. VAT)",  grand_after_disc, C_SUMMARY, True,  False, C_TEXT),
                            (f"VAT {vat:.0f}%",   grand_vat,        C_SUMMARY, False, False, C_MUTED),
                            ("TOTAL (INC. VAT)",  grand_total,      C_TOTAL,   True,  False, C_TOTAL_TXT),
                        ]
                        for label,vals,sbg,sbold,is_disc,stxt in sum_defs:
                            ws.merge_cells(start_row=current_row, start_column=1,
                                           end_row=current_row, end_column=4)
                            cs(current_row, 1, label, bold=sbold, align="right",
                               bg=sbg, size=9, color=stxt)
                            for si, shop in enumerate(shops):
                                col_u = SHOP_START + si*2
                                col_t = col_u + 1
                                v     = vals.get(shop,0)
                                is_b  = label=="TOTAL (INC. VAT)" and si==best_i
                                cell_bg  = C_WIN   if is_b else sbg
                                cell_txt = C_WIN_TXT if is_b else stxt
                                disp  = -v if is_disc else v
                                ws.merge_cells(start_row=current_row, start_column=col_u,
                                               end_row=current_row, end_column=col_t)
                                cs(current_row, col_u, disp, bold=sbold or is_b, align="right",
                                   bg=cell_bg, fmt='#,##0.00', size=10, color=cell_txt)
                            ws.row_dimensions[current_row].height = 18
                            current_row += 1

                        current_row += 2  # ช่องว่างระหว่างกลุ่ม

                    # สรุปรวมทุกกลุ่ม (ถ้ามีหลายกลุ่ม)
                    if len([g for g in groups if g.get("items")]) > 1:
                        ws.merge_cells(start_row=current_row, start_column=1,
                                       end_row=current_row, end_column=total_cols)
                        sg = ws.cell(row=current_row, column=1, value="สรุปรวมทุกกลุ่ม  —  TOTAL (INC. VAT)")
                        sg.font      = Font(bold=True, size=11, color=C_WHITE, name="Calibri")
                        sg.fill      = PatternFill("solid", fgColor="0F6E56")
                        sg.alignment = Alignment(horizontal="left", vertical="center", indent=1)
                        sg.border    = bdr_none
                        ws.row_dimensions[current_row].height = 22
                        current_row += 1

                        valid_all2   = {s:all_totals[s] for s in shops if all_totals[s]>0}
                        best_all_i   = shops.index(min(valid_all2,key=valid_all2.get)) if valid_all2 else None
                        ws.merge_cells(start_row=current_row, start_column=1,
                                       end_row=current_row, end_column=4)
                        cs(current_row, 1, "", bg=C_TOTAL)
                        for si, shop in enumerate(shops):
                            col_u = SHOP_START + si*2
                            col_t = col_u + 1
                            v     = all_totals[shop]
                            is_b  = si == best_all_i
                            ws.merge_cells(start_row=current_row, start_column=col_u,
                                           end_row=current_row, end_column=col_t)
                            cs(current_row, col_u, v, bold=is_b, align="right",
                               bg=C_WIN if is_b else C_TOTAL,
                               fmt='#,##0.00', size=12,
                               color=C_WIN_TXT if is_b else C_TOTAL_TXT)
                        ws.row_dimensions[current_row].height = 24

                    ws.freeze_panes = "E4"
                    out = io.BytesIO(); wb.save(out); out.seek(0)
                    return out.getvalue()

                # PDF Export
                def export_pdf():
                    from reportlab.lib.pagesizes import A4, landscape
                    from reportlab.lib import colors
                    from reportlab.lib.styles import ParagraphStyle
                    from reportlab.lib.units import mm
                    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                                     Paragraph, Spacer)
                    from reportlab.pdfbase import pdfmetrics
                    from reportlab.pdfbase.ttfonts import TTFont
                    import urllib.request, os

                    font_path = "/tmp/Sarabun-Regular.ttf"
                    font_b_path = "/tmp/Sarabun-Bold.ttf"
                    if not os.path.exists(font_path):
                        urllib.request.urlretrieve(
                            "https://github.com/google/fonts/raw/main/ofl/sarabun/Sarabun-Regular.ttf", font_path)
                    if not os.path.exists(font_b_path):
                        urllib.request.urlretrieve(
                            "https://github.com/google/fonts/raw/main/ofl/sarabun/Sarabun-Bold.ttf", font_b_path)
                    try:
                        pdfmetrics.registerFont(TTFont("TH", font_path))
                        pdfmetrics.registerFont(TTFont("TH-B", font_b_path))
                        F, FB = "TH", "TH-B"
                    except:
                        F = FB = "Helvetica"

                    buf = io.BytesIO()
                    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                                            leftMargin=15*mm, rightMargin=15*mm,
                                            topMargin=15*mm, bottomMargin=15*mm)
                    story = []

                    def p(txt, bold=False, size=10, color=colors.HexColor("#1E293B"), align="LEFT"):
                        s = ParagraphStyle("s", fontName=FB if bold else F, fontSize=size,
                                           textColor=color,
                                           alignment={"LEFT":0,"CENTER":1,"RIGHT":2}.get(align,0),
                                           leading=size*1.3)
                        return Paragraph(str(txt or ""), s)

                    GREEN  = colors.HexColor("#1E293B")
                    WIN_BG = colors.HexColor("#DCFCE7")
                    WIN_TX = colors.HexColor("#166534")
                    DISC_B = colors.HexColor("#FEF9C3")
                    DISC_T = colors.HexColor("#854D0E")
                    SUM_B  = colors.HexColor("#F8FAFC")
                    TOT_B  = colors.HexColor("#E1F5EE")
                    TOT_T  = colors.HexColor("#0F6E56")
                    WHT    = colors.white
                    MUT    = colors.HexColor("#64748B")
                    shop_hdr_colors = [colors.HexColor(h) for h in
                                       ["#334155","#334155","#334155","#334155","#334155"]]

                    story.append(p(f"เปรียบเทียบราคา — {st.session_state['project_name']}",
                                   bold=True, size=15, color=GREEN))
                    story.append(p(f"วันที่ {doc_date.strftime('%d/%m/%Y')}   |   VAT {vat:.0f}%   |   {st.session_state['doc_title']}",
                                   size=9, color=MUT))
                    story.append(Spacer(1,6*mm))

                    page_w = landscape(A4)[0] - 30*mm
                    fix_w  = page_w * 0.32
                    shop_w = (page_w - fix_w) / len(shops)
                    col_ws = [fix_w*0.07, fix_w*0.45, fix_w*0.10, fix_w*0.10] + \
                             [shop_w*0.5, shop_w*0.5] * len(shops)

                    for g_idx, grp in enumerate(groups):
                        items_data = grp.get("items",[]); shop_disc = grp.get("discounts",{})
                        if not items_data: continue
                        grand_sub,grand_disc_g,grand_after_disc,grand_vat,grand_total = calc_group(
                            items_data, shops, vat, shop_disc)
                        valid  = {s:grand_total[s] for s in shops if grand_sub[s]>0}
                        best_s = min(valid,key=valid.get) if valid else None

                        tdata = []
                        cmds  = [("GRID",(0,0),(-1,-1),0.5,colors.HexColor("#E2E8F0")),
                                 ("FONTSIZE",(0,0),(-1,-1),9),
                                 ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
                                 ("TOPPADDING",(0,0),(-1,-1),4),
                                 ("BOTTOMPADDING",(0,0),(-1,-1),4)]

                        # group header
                        tdata.append([p(f"กลุ่ม {g_idx+1}  :  {grp['name']}",bold=True,size=10,color=WHT)]+[""]*(3+len(shops)*2))
                        cmds += [("SPAN",(0,0),(-1,0)),
                                 ("BACKGROUND",(0,0),(-1,0),GREEN),
                                 ("ROWBACKGROUNDS",(0,0),(-1,0),[GREEN])]

                        # col headers
                        h1 = [p("#",bold=True,color=WHT,align="CENTER"),
                               p("รายการ",bold=True,color=WHT),
                               p("จำนวน",bold=True,color=WHT,align="CENTER"),
                               p("หน่วย",bold=True,color=WHT,align="CENTER")]
                        h2 = ["","","",""]
                        for si,s in enumerate(shops):
                            lbl = ("★ " if s==best_s else "") + s
                            h1 += [p(lbl,bold=True,color=WHT,align="CENTER"), ""]
                            h2 += [p("Unit Price",bold=False,size=8,color=colors.HexColor("#CBD5E1"),align="CENTER"),
                                   p("Total",bold=False,size=8,color=colors.HexColor("#CBD5E1"),align="CENTER")]
                            cmds += [("SPAN",(4+si*2,1),(5+si*2,1)),
                                     ("BACKGROUND",(4+si*2,1),(5+si*2,2),
                                      colors.HexColor("#1E293B") if s==best_s else colors.HexColor("#334155"))]
                        tdata += [h1, h2]
                        for c in range(4):
                            cmds.append(("SPAN",(c,1),(c,2)))
                        cmds += [("BACKGROUND",(0,1),(-1,2),colors.HexColor("#334155")),
                                 ("SPAN",(0,0),(-1,0))]

                        row_i = 3
                        for i, it in enumerate(items_data):
                            qty = float(it["qty"])
                            raw_tots  = [float(it["prices"].get(s,0))*qty for s in shops]
                            disc_tots = [float(it.get("item_discounts",{}).get(s,0))*qty for s in shops]
                            after_t   = [max(raw_tots[si]-disc_tots[si],0) for si in range(len(shops))]
                            valid_tv  = [v for v in after_t if v>0]
                            best_p    = min(valid_tv) if valid_tv else None
                            row_bg    = WHT if i%2==0 else SUM_B

                            dr = [p(str(i+1),align="CENTER"),
                                  p(f"[{it.get('category','')}] {it['name']}",bold=True),
                                  p(str(int(qty)),align="CENTER"),
                                  p(it["unit"],align="CENTER")]
                            for si,s in enumerate(shops):
                                price = float(it["prices"].get(s,0))
                                raw_t = raw_tots[si]
                                is_b  = best_p and after_t[si]==best_p and after_t[si]>0
                                c_bg  = WIN_BG if is_b else row_bg
                                c_tx  = WIN_TX if is_b else colors.HexColor("#1E293B")
                                dr += [p(f"฿{price:,.2f}",align="RIGHT",color=c_tx),
                                       p(f"฿{raw_t:,.2f}",bold=is_b,align="RIGHT",color=c_tx)]
                                cmds.append(("BACKGROUND",(4+si*2,row_i),(5+si*2,row_i),c_bg))
                            tdata.append(dr)
                            cmds.append(("BACKGROUND",(0,row_i),(3,row_i),row_bg))
                            row_i += 1

                            has_disc = any(float(it.get("item_discounts",{}).get(s,0))>0 for s in shops)
                            if has_disc:
                                drow = ["","","",p(f"ส่วนลด {it['name']}",size=8,color=DISC_T,align="RIGHT")]
                                for si,s in enumerate(shops):
                                    dt = disc_tots[si]
                                    drow += ["",p(f"-฿{dt:,.2f}" if dt>0 else "",size=8,color=DISC_T,align="RIGHT")]
                                    cmds.append(("BACKGROUND",(4+si*2,row_i),(5+si*2,row_i),DISC_B))
                                tdata.append(drow); cmds.append(("BACKGROUND",(0,row_i),(3,row_i),DISC_B))
                                row_i += 1

                        for label,vals,sbg,sbold,is_disc,stx in [
                            ("SPECIAL DISCOUNT",  grand_disc_g,    DISC_B, False, True,  DISC_T),
                            ("TOTAL (EXC. VAT)",  grand_after_disc, SUM_B, True,  False, colors.HexColor("#1E293B")),
                            (f"VAT {vat:.0f}%",   grand_vat,        SUM_B, False, False, MUT),
                            ("TOTAL (INC. VAT)",  grand_total,      TOT_B, True,  False, TOT_T),
                        ]:
                            srow = ["","","",p(label,bold=sbold,align="RIGHT",color=stx)]
                            for si,s in enumerate(shops):
                                v    = vals.get(s,0)
                                is_b = label=="TOTAL (INC. VAT)" and s==best_s
                                c_bg = WIN_BG if is_b else sbg
                                c_tx = WIN_TX if is_b else stx
                                disp = f"-฿{v:,.2f}" if is_disc else f"฿{v:,.2f}"
                                srow += ["",p(disp,bold=sbold or is_b,align="RIGHT",color=c_tx)]
                                cmds.append(("BACKGROUND",(4+si*2,row_i),(5+si*2,row_i),c_bg))
                            tdata.append(srow)
                            cmds += [("BACKGROUND",(0,row_i),(3,row_i),sbg),
                                     ("SPAN",(0,row_i),(3,row_i))]
                            for si in range(len(shops)):
                                cmds.append(("SPAN",(4+si*2,row_i),(5+si*2,row_i)))
                            row_i += 1

                        t = Table(tdata, colWidths=col_ws, repeatRows=3)
                        t.setStyle(TableStyle(cmds))
                        story.append(t)
                        story.append(Spacer(1,8*mm))

                    doc.build(story)
                    buf.seek(0)
                    return buf.getvalue()

                col_xl, col_pdf = st.columns(2)
                with col_xl:
                    st.download_button("📊 ดาวน์โหลด Excel",
                        data=export_excel(),
                        file_name=f"{st.session_state['doc_title']}_{doc_date.strftime('%Y%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True)
                with col_pdf:
                    try:
                        st.download_button("📄 ดาวน์โหลด PDF",
                            data=export_pdf(),
                            file_name=f"{st.session_state['doc_title']}_{doc_date.strftime('%Y%m%d')}.pdf",
                            mime="application/pdf",
                            use_container_width=True)
                    except Exception as e:
                        st.error(f"สร้าง PDF ไม่สำเร็จ: {e}")

# ============================================================
# MENU: ร้านค้า
# ============================================================
elif "🏪" in menu:
    st.markdown("""
    <div class="page-header">
      <div class="page-header-icon">🏪</div>
      <div>
        <div class="page-header-title">ฐานข้อมูลร้านค้า</div>
        <div class="page-header-sub">บันทึกข้อมูลร้านค้าที่ใช้บ่อย เรียกใช้ในโครงการได้เลย</div>
      </div>
    </div>
    """, unsafe_allow_html=True)

    with st.expander("➕ เพิ่มร้านค้าใหม่", expanded=False):
        with st.form("add_shop_form"):
            sc1,sc2 = st.columns(2)
            shop_name    = sc1.text_input("ชื่อบริษัท / ร้านค้า *")
            shop_contact = sc2.text_input("ชื่อผู้ติดต่อ")
            sc3,sc4 = st.columns(2)
            shop_phone = sc3.text_input("เบอร์โทร")
            shop_email = sc4.text_input("อีเมล")
            shop_address = st.text_area("ที่อยู่", height=60)
            sc5,sc6 = st.columns(2)
            shop_payment = sc5.selectbox("เงื่อนไขการชำระ",
                                         ["เงินสด","โอน","เครดิต 30 วัน","เครดิต 60 วัน","อื่นๆ"])
            shop_notes = sc6.text_input("หมายเหตุ")
            if st.form_submit_button("บันทึกร้านค้า", type="primary"):
                if shop_name:
                    save_shop_db({"name":shop_name,"contact":shop_contact,"phone":shop_phone,
                                  "email":shop_email,"address":shop_address,
                                  "payment_terms":shop_payment,"notes":shop_notes},
                                 st.session_state["display_name"])
                    st.success(f"เพิ่ม '{shop_name}' แล้วครับ!"); st.rerun()
                else:
                    st.error("กรุณาใส่ชื่อร้านค้า")

    shops_db = load_shops_db()
    if not shops_db:
        st.markdown("""<div style="text-align:center;padding:3rem;color:#94a3b8">
          <div style="font-size:36px">🏪</div><div style="margin-top:8px">ยังไม่มีร้านค้า</div></div>""",
                    unsafe_allow_html=True)
    else:
        search_shop = st.text_input("", placeholder="🔍 ค้นหาร้านค้า...", label_visibility="collapsed")
        if search_shop:
            shops_db = [s for s in shops_db if search_shop.lower() in s.get("name","").lower()
                        or search_shop in s.get("phone","")]
        for s in shops_db:
            sid = s.get("id") or s.get("ID","")
            with st.expander(f"**{s.get('name','')}**  ·  {s.get('contact','')}  ·  {s.get('phone','')}", expanded=False):
                c1,c2,c3,c4 = st.columns(4)
                c1.markdown(f"📞 {s.get('phone','—')}")
                c2.markdown(f"✉️ {s.get('email','—')}")
                c3.markdown(f"💳 {s.get('payment_terms','—')}")
                c4.markdown(f"📝 {s.get('notes','—')}")
                if s.get("address"):
                    st.markdown(f"📍 {s.get('address')}")
                if sid and st.button("🗑 ลบร้านค้านี้", key=f"dshop_{sid}"):
                    delete_shop_db(sid); st.rerun()

# ============================================================
# MENU: Templates
# ============================================================
elif "📦" in menu:
    st.markdown("""
    <div class="page-header">
      <div class="page-header-icon">📦</div>
      <div>
        <div class="page-header-title">Template สินค้า</div>
        <div class="page-header-sub">บันทึกรายการสินค้าที่ใช้บ่อย เรียกใช้ในโครงการได้เลย</div>
      </div>
    </div>
    """, unsafe_allow_html=True)

    with st.expander("➕ สร้าง Template ใหม่", expanded=False):
        tc1,tc2 = st.columns(2)
        tmpl_name = tc1.text_input("ชื่อ Template *", key="new_tmpl_name")
        tmpl_cat  = tc2.selectbox("หมวดหมู่",
                                   ["ทั่วไป","ก่อสร้าง","IT/Network","ไฟฟ้า","ประปา","อื่นๆ"],
                                   key="new_tmpl_cat")

        if "tmpl_draft_items" not in st.session_state:
            st.session_state["tmpl_draft_items"] = [{"name":"","unit":"set","qty":1.0,"category":"วัสดุ/อุปกรณ์"}]

        to_del_tmpl = None
        for ti, dit in enumerate(st.session_state["tmpl_draft_items"]):
            ta,tb,tc_col,td,te = st.columns([3,1,1,1.5,0.5])
            dit["name"]     = ta.text_input("ชื่อสินค้า",   dit["name"],     key=f"ti_name_{ti}")
            dit["unit"]     = tb.text_input("หน่วย",        dit["unit"],     key=f"ti_unit_{ti}")
            dit["qty"]      = tc_col.number_input("จำนวน",   value=float(dit["qty"]), min_value=0.0, step=1.0, key=f"ti_qty_{ti}")
            dit["category"] = td.selectbox("ประเภท", ITEM_CATEGORIES,
                                            index=ITEM_CATEGORIES.index(dit.get("category","วัสดุ/อุปกรณ์"))
                                            if dit.get("category") in ITEM_CATEGORIES else 0,
                                            key=f"ti_cat_{ti}")
            if te.button("🗑", key=f"ti_del_{ti}") and len(st.session_state["tmpl_draft_items"]) > 1:
                to_del_tmpl = ti

        if to_del_tmpl is not None:
            st.session_state["tmpl_draft_items"].pop(to_del_tmpl); st.rerun()

        col_add, col_save = st.columns([1,2])
        if col_add.button("➕ เพิ่มรายการ"):
            st.session_state["tmpl_draft_items"].append({"name":"","unit":"set","qty":1.0,"category":"วัสดุ/อุปกรณ์"})
            st.rerun()
        if col_save.button("💾 บันทึก Template", type="primary"):
            if tmpl_name:
                valid_items = [it for it in st.session_state["tmpl_draft_items"] if it["name"].strip()]
                if valid_items:
                    save_template({"name":tmpl_name,"category":tmpl_cat,"items":valid_items},
                                  st.session_state["display_name"])
                    st.session_state["tmpl_draft_items"] = [{"name":"","unit":"set","qty":1.0,"category":"วัสดุ/อุปกรณ์"}]
                    st.success(f"บันทึก Template '{tmpl_name}' แล้วครับ!"); st.rerun()
                else: st.error("กรุณาใส่รายการสินค้าอย่างน้อย 1 รายการ")
            else: st.error("กรุณาใส่ชื่อ Template")

    templates = load_templates()
    if not templates:
        st.markdown("""<div style="text-align:center;padding:3rem;color:#94a3b8">
          <div style="font-size:36px">📦</div><div style="margin-top:8px">ยังไม่มี Template</div></div>""",
                    unsafe_allow_html=True)
    else:
        for t in templates:
            try: items_list = json.loads(t.get("items","[]"))
            except: items_list = []
            tid = t.get("id") or t.get("ID","")
            with st.expander(f"**{t.get('name','')}**  [{t.get('category','')}]  —  {len(items_list)} รายการ", expanded=False):
                for it in items_list:
                    st.markdown(f"- **{it.get('name','')}** &nbsp; {it.get('unit','')} × {it.get('qty','')} &nbsp; [{it.get('category','')}]")
                st.caption(f"สร้างโดย {t.get('created_by','')}  ·  {t.get('created_at','')}")
                if tid and st.button("🗑 ลบ Template", key=f"dtmpl_{tid}"):
                    delete_template(tid); st.rerun()
